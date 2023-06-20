# encoding:utf-8
import json
import re
import time
import numpy as np
import openai
import openai.error
import requests
from bot.bot import Bot
from bot.chatgpt.chat_gpt_session import ChatGPTSession
from bot.openai.open_ai_image import OpenAIImage
from bot.session_manager import SessionManager
from bridge.context import ContextType
from bridge.reply import Reply, ReplyType
from common.log import logger
from common.token_bucket import TokenBucket
from config import conf, load_config
import openpyxl

def xingpan(socket_birth_data: dict):
    header = {"Content-Type": "application/json;charset=UTF-8"}
    natal_url = 'http://www.xingpan.vip/astrology/chart/natal'
    natal_required_datas = {
        "access_token": "73160ff79f74748e589b0f4e9f04097b",
        "h_sys": "P",
        "planets": ["0", "1", "2", "3", "4", "5", "6", "7", "8", "9", "t","H"],
        "tz": "8.00",
        list(socket_birth_data.keys())[0]: list(socket_birth_data.values())[0],
        list(socket_birth_data.keys())[1]: list(socket_birth_data.values())[1],
        list(socket_birth_data.keys())[2]: list(socket_birth_data.values())[2],
        'is_corpus': 1,
        'svg_type': -1,
        'tomorrow_type': 1,
        'phase': {'0': 0.5, '180': 6, '120': 6, '90': 6, '60': 6}
    }
    returned_natal = requests.post(natal_url, headers=header, data=json.dumps(natal_required_datas)).json()  # 得到星盘数据
    house_and_sign = []  # 宫位和落座
    xiangwei = []  # 相位
    natal_text = ''
    for i in range(len(returned_natal['data']['planet'])):
        house_and_sign.append({'type': 4,
                               'planet_id': returned_natal['data']['planet'][i]['code_name'],
                               'house_id': returned_natal['data']['planet'][i]['house_id'], })
        house_and_sign.append({'type': 5,
                               'planet_id': returned_natal['data']['planet'][i]['code_name'],
                               'sign_id': returned_natal['data']['planet'][i]['sign']['sign_id'], })  # 从星盘里获取宫位和落座的数据
        for j in range(len(returned_natal['data']['planet'][i]['planet_allow_degree'])):
            xiangwei.append({'type': 6,
                             'planet_id1': returned_natal['data']['planet'][i]['code_name'],
                             'planet_id2': returned_natal['data']['planet'][i]['planet_allow_degree'][j]['code_name'],
                             'degree': returned_natal['data']['planet'][i]['planet_allow_degree'][j][
                                 'allow']})  # 从星盘里获取相位的数据
    socket_required_hs = {'chartType': 'natal',
                          "access_token": "73160ff79f74748e589b0f4e9f04097b",
                          "fallInto": json.dumps(house_and_sign)}  # 转化宫位和落座数据以传入接口
    socket_required_xw = {'chartType': 'natal',
                          "access_token": "73160ff79f74748e589b0f4e9f04097b",
                          "fallInto": json.dumps(xiangwei)}  # 转化相位数据以传入接口
    yuliaourl = 'http://www.xingpan.vip/astrology/corpusconstellation/getlist'
    hs_text = requests.post(yuliaourl, headers=header, data=json.dumps(socket_required_hs)).json()  # 宫位和落座语料
    for h in range(len(hs_text['data'])):
        natal_text += hs_text['data'][h]['title'] + '，'
    xw_text = requests.post(yuliaourl, headers=header, data=json.dumps(socket_required_xw)).json()  # 相位语料
    for x in range(len(xw_text['data'])):
        natal_text += xw_text['data'][x]['title'] + '，'
    natal_text = '['+natal_text[:-1] + ']'  # 结尾改成句号
    return returned_natal, natal_text
def get_title(type: int, id1, id2, agree=999):
    house = ['', '1宫', '2宫', '3宫', '4宫', '5宫', '6宫', '7宫', '8宫', '9宫', '10宫', '11宫', '12宫']
    sign = ['白羊', '金牛', '双子', '巨蟹', '狮子', '处女', '天秤', '天蝎', '射手', '摩羯',
            '水瓶', '双鱼']
    planet = {'0': '太阳', '1': '月亮', '2': '水星', '3': '金星', '4': '火星', '5': '木星', '6': '土星', '7': '天王',
              '8': '海王', '9': '冥王', 't': '北交', 'H': '婚神'}
    xiagnwei = {0: '合', 180: '冲', 120: '拱', 90: '刑', 60: '六合'}
    title = ''
    if type == 3:
        title = house[id1]+sign[id2]   # 宫位星座
    elif type == 4:
        title = planet[id1]+house[id2]  # 行星宫位
    elif type == 5:
        title = planet[id1]+sign[id2]  # 行星星座
    elif type == 6:
        if agree == 80 or agree == 180:
            title = planet[id1] + planet[id2] + '正'
        else:
            title = planet[id1] + planet[id2] + '负'  # 行星相位
    return title
def good_or_bad(score):
    if score=='0P':
        score=0
    score = int(score)
    if score < 1:
        return '差'
    elif score < 5:
        return '中'
    else:
        return '好'
def band_or_not(question1:str,question2:str):
    relevance = openai.Completion.create(
        api_key=conf().get("open_ai_api_key"),
        model="text-davinci-003",
        prompt="判断两个问题之间有没有关联，如果有回复1，没有回复0，问题1："+question1+"问题2："+question2,
        max_tokens=256,
        temperature=0.7
    )
    return relevance.choices[0]["text"].strip()
def true_love(white: int, sign_guardian):
    true_love_time = []
    true_love_night = [
    # {'planet': 'Moon', 'age': 0, 'day': 0},
    # {'planet': 'Saturn', 'age': 1, 'day': 104},
    # {'planet': 'Jupiter', 'age': 2, 'day': 209},
    # {'planet': 'Mars', 'age': 3, 'day': 313},
    # {'planet': 'Sun', 'age': 5, 'day': 52},
    # {'planet': 'Venus', 'age': 6, 'day': 156},
    # {'planet': 'Mercury', 'age': 7, 'day': 261},
    # {'planet': 'Saturn', 'age': 9, 'day': 0},
    # {'planet': 'Jupiter', 'age': 10, 'day': 209},
    # {'planet': 'Mars', 'age': 12, 'day': 52},
    # {'planet': 'Sun', 'age': 13, 'day': 261},
    # {'planet': 'Venus', 'age': 15, 'day': 104},
    # {'planet': 'Mercury', 'age': 16, 'day': 313},
    {'planet': 'Moon', 'age': 18, 'day': 156},
    {'planet': 'Jupiter', 'age': 20, 'day': 0},
    {'planet': 'Mars', 'age': 21, 'day': 261},
    {'planet': 'Sun', 'age': 23, 'day': 156},
    {'planet': 'Venus', 'age': 25, 'day': 52},
    {'planet': 'Mercury', 'age': 26, 'day': 313},
    {'planet': 'Moon', 'age': 28, 'day': 209},
    {'planet': 'Saturn', 'age': 30, 'day': 104},
    {'planet': 'Mars', 'age': 32, 'day': 0},
    {'planet': 'Sun', 'age': 33, 'day': 0},
    {'planet': 'Venus', 'age': 34, 'day': 0},
    {'planet': 'Mercury', 'age': 35, 'day': 0},
    {'planet': 'Moon', 'age': 36, 'day': 0},
    {'planet': 'Saturn', 'age': 37, 'day': 0},
    {'planet': 'Jupiter', 'age': 38, 'day': 0},
    {'planet': 'Sun', 'age': 39, 'day': 0},
    {'planet': 'Venus', 'age': 40, 'day': 156},
    {'planet': 'Mercury', 'age': 41, 'day': 313},
    {'planet': 'Moon', 'age': 43, 'day': 104},
    {'planet': 'Saturn', 'age': 44, 'day': 261},
    {'planet': 'Jupiter', 'age': 46, 'day': 52},
    {'planet': 'Mars', 'age': 47, 'day': 209},
    {'planet': 'Venus', 'age': 49, 'day': 0},
    {'planet': 'Mercury', 'age': 50, 'day': 52},
    {'planet': 'Moon', 'age': 51, 'day': 104},
    {'planet': 'Saturn', 'age': 52, 'day': 156},
    {'planet': 'Jupiter', 'age': 53, 'day': 209},
    {'planet': 'Mars', 'age': 54, 'day': 261},
    {'planet': 'Sun', 'age': 55, 'day': 313},
    {'planet': 'Mercury', 'age': 57, 'day': 0},
    {'planet': 'Moon', 'age': 58, 'day': 313},
    {'planet': 'Saturn', 'age': 60, 'day': 261},
    {'planet': 'Jupiter', 'age': 62, 'day': 209},
    {'planet': 'Mars', 'age': 64, 'day': 156},
    {'planet': 'Sun', 'age': 66, 'day': 104},
    {'planet': 'Venus', 'age': 68, 'day': 52},
    {'planet': 'meanNode', 'age': 70, 'day': 0},
    {'planet': 'MeanSouthNode', 'age': 73, 'day': 0}
]
    true_love_day = [
                     # {'planet': 'Sun', 'age': 0, 'day': 0},
                     # {'planet': 'Venus', 'age': 1, 'day': 156},
                     # {'planet': 'Mercury', 'age': 2, 'day': 313},
                     # {'planet': 'Moon', 'age': 4, 'day': 104},
                     # {'planet': 'Saturn', 'age': 5, 'day': 261},
                     # {'planet': 'Jupiter', 'age': 7, 'day': 52},
                     # {'planet': 'Mars', 'age': 8, 'day': 209},
                     # {'planet': 'Venus', 'age': 10, 'day': 0},
                     # {'planet': 'Mercury', 'age': 11, 'day': 52},
                     # {'planet': 'Moon', 'age': 12, 'day': 104},
                     # {'planet': 'Saturn', 'age': 13, 'day': 156},
                     # {'planet': 'Jupiter', 'age': 14, 'day': 209},
                     # {'planet': 'Mars', 'age': 15, 'day': 261},
                     # {'planet': 'Sun', 'age': 16, 'day': 313},
                     {'planet': 'Mercury', 'age': 18, 'day': 0},
                     {'planet': 'Moon', 'age': 19, 'day': 313},
                     {'planet': 'Saturn', 'age': 21, 'day': 261},
                     {'planet': 'Jupiter', 'age': 23, 'day': 209},
                     {'planet': 'Mars', 'age': 25, 'day': 156},
                     {'planet': 'Sun', 'age': 27, 'day': 104},
                     {'planet': 'Venus', 'age': 29, 'day': 52},
                     {'planet': 'Moon', 'age': 31, 'day': 0},
                     {'planet': 'Saturn', 'age': 32, 'day': 104},
                     {'planet': 'Jupiter', 'age': 33, 'day': 209},
                     {'planet': 'Mars', 'age': 34, 'day': 313},
                     {'planet': 'Sun', 'age': 36, 'day': 52},
                     {'planet': 'Venus', 'age': 37, 'day': 156},
                     {'planet': 'Mercury', 'age': 38, 'day': 261},
                     {'planet': 'Saturn', 'age': 40, 'day': 0},
                     {'planet': 'Jupiter', 'age': 41, 'day': 209},
                     {'planet': 'Mars', 'age': 43, 'day': 52},
                     {'planet': 'Sun', 'age': 44, 'day': 261},
                     {'planet': 'Venus', 'age': 46, 'day': 104},
                     {'planet': 'Mercury', 'age': 47, 'day': 313},
                     {'planet': 'Moon', 'age': 49, 'day': 156},
                     {'planet': 'Jupiter', 'age': 51, 'day': 0},
                     {'planet': 'Mars', 'age': 52, 'day': 261},
                     {'planet': 'Sun', 'age': 54, 'day': 156},
                     {'planet': 'Venus', 'age': 56, 'day': 52},
                     {'planet': 'Mercury', 'age': 57, 'day': 313},
                     {'planet': 'Moon', 'age': 59, 'day': 209},
                     {'planet': 'Saturn', 'age': 61, 'day': 104},
                     {'planet': 'Mars', 'age': 63, 'day': 0},
                     {'planet': 'Sun', 'age': 64, 'day': 0},
                     {'planet': 'Venus', 'age': 65, 'day': 0},
                     {'planet': 'Mercury', 'age': 66, 'day': 0},
                     {'planet': 'Moon', 'age': 67, 'day': 0},
                     {'planet': 'Saturn', 'age': 68, 'day': 0},
                     {'planet': 'Jupiter', 'age': 69, 'day': 0},
                     {'planet': 'meanNode', 'age': 70, 'day': 0},
                     {'planet': 'MeanSouthNode', 'age': 73, 'day': 0}]
    if white == 0:
        for i in true_love_day:
            if sign_guardian == i['planet']:
                true_love_time.append(str(i['age'])+'岁')
    elif white == 1:
        for i in true_love_night:
            if sign_guardian == i['planet']:
                true_love_time.append(str(i['age'])+'岁')
    return str(true_love_time)[3:-2]

# OpenAI对话模型API (可用)
class ChatGPTBot(Bot, OpenAIImage):
    def __init__(self):
        super().__init__()
        # set the default api_key
        openai.api_key = conf().get("open_ai_api_key")
        if conf().get("open_ai_api_base"):
            openai.api_base = conf().get("open_ai_api_base")
        proxy = conf().get("proxy")
        if proxy:
            openai.proxy = proxy
        if conf().get("rate_limit_chatgpt"):
            self.tb4chatgpt = TokenBucket(conf().get("rate_limit_chatgpt", 20))

        self.sessions = SessionManager(
            ChatGPTSession, model=conf().get("model") or "gpt-3.5-turbo"
        )
        self.args = {
            "model": conf().get("model") or "gpt-3.5-turbo",  # 对话模型的名称
            "temperature": conf().get("temperature", 0.7),  # 值在[0,1]之间，越大表示回复越具有不确定性
            "max_tokens": 1024,  # 回复最大的字符数
            "top_p": 1,
            "frequency_penalty": conf().get(
                "frequency_penalty", 0.0
            ),  # [-2,2]之间，该值越大则更倾向于产生不同的内容
            "presence_penalty": conf().get(
                "presence_penalty", 0.0
            ),  # [-2,2]之间，该值越大则更倾向于产生不同的内容
            "request_timeout": conf().get(
                "request_timeout", None
            ),  # 请求超时时间，openai接口默认设置为600，对于难问题一般需要较长时间
            "timeout": conf().get("request_timeout", None),  # 重试超时时间，在这个时间内，将会自动重试
        }

    def reply(self, query, context=None):
        # acquire reply content
        if context.type == ContextType.TEXT:
            logger.info("[CHATGPT] query={}".format(query))

            session_id = context["session_id"]
            reply = None
            clear_memory_commands = conf().get("clear_memory_commands", ["#清除记忆"])
            if query in clear_memory_commands:
                self.sessions.clear_session(session_id)
                reply = Reply(ReplyType.INFO, "记忆已清除")
            elif query == "#清除所有":
                self.sessions.clear_all_session()
                reply = Reply(ReplyType.INFO, "所有人记忆已清除")
            elif query == "#更新配置":
                load_config()
                reply = Reply(ReplyType.INFO, "配置已更新")
            if reply:
                return reply
            session = self.sessions.session_query(query, session_id)
            logger.debug("[CHATGPT] session query={}".format(session.messages))

            api_key = context.get("openai_api_key")

            # if context.get('stream'):
            #     # reply in stream
            #     return self.reply_text_stream(query, new_query, session_id)

            reply_content = self.reply_text(session, api_key)
            logger.debug(
                "[CHATGPT] new_query={}, session_id={}, reply_cont={}, completion_tokens={}".format(
                    session.messages,
                    session_id,
                    reply_content["content"],
                    reply_content["completion_tokens"],
                )
            )
            if (
                    reply_content["completion_tokens"] == 0
                    and len(reply_content["content"]) > 0
            ):
                reply = Reply(ReplyType.ERROR, reply_content["content"])
            elif reply_content["completion_tokens"] > 0:
                self.sessions.session_reply(
                    reply_content["content"], session_id, reply_content["total_tokens"]
                )
                reply = Reply(ReplyType.TEXT, reply_content["content"])
            else:
                reply = Reply(ReplyType.ERROR, reply_content["content"])
                logger.debug("[CHATGPT] reply {} used 0 tokens.".format(reply_content))
            return reply

        elif context.type == ContextType.IMAGE_CREATE:
            ok, retstring = self.create_img(query, 0)
            reply = None
            if ok:
                reply = Reply(ReplyType.IMAGE_URL, retstring)
            else:
                reply = Reply(ReplyType.ERROR, retstring)
            return reply
        else:
            reply = Reply(ReplyType.ERROR, "Bot不支持处理{}类型的消息".format(context.type))
            return reply

    def reply_text(self, session: ChatGPTSession, api_key=None, retry_count=0) -> dict:
        """
        call openai's ChatCompletion to get the answer
        :param session: a conversation session
        :param session_id: session id
        :param retry_count: retry count
        :return: {}
        """
        try:
            if conf().get("rate_limit_chatgpt") and not self.tb4chatgpt.get_token():
                raise openai.error.RateLimitError("RateLimitError: rate limit exceeded")
            # if api_key == None, the default openai.api_key will be used
            user_question = session.messages[-1]['content']
            if session.natal == {}:
                birthdata_transform = openai.ChatCompletion.create(
                    api_key=api_key,
                    messages=[
                        {"role": "system",
                         "content": "将我输入的生日和出生地相关信息按照以下格式返回，{\"birthday\": \"1994-09-0115 5:30\",\"latitude\": \"30.60\",\"longitude\": \"114.30\"}，城市转化为对应经纬度，如果没有输入时间默认为中午12点。"},
                        {"role": "user", "content": user_question}],
                    **self.args)

                check=birthdata_transform.choices[0]["message"]["content"]
                if "longitude" in check and "birthday" in check and "latitude" in check:
                    user_birth_data = json.loads(re.findall(r'{.*}', check)[0])
                    session.natal, session.natal_titles = xingpan(user_birth_data)
                    session.reset()
                    #session.set_system_prompt('你是一个占星师，并且相信占星学是真理，你需要根据占星学回答我的问题。')
                    session.add_reply(session.natal_titles)
                    return {
                        "total_tokens": birthdata_transform["usage"]["total_tokens"],
                        "completion_tokens": birthdata_transform["usage"]["completion_tokens"],
                        "content": '现在我已经获得您的星盘，请问您想占星什么问题？'
                    }
                # logger.info("[ChatGPT] reply={}, total_tokens={}".format(response.choices[0]['message']['content'], response["usage"]["total_tokens"]))
                session.reset()
                return {
                    "total_tokens": birthdata_transform["usage"]["total_tokens"],
                    "completion_tokens": birthdata_transform["usage"]["completion_tokens"],
                    "content": '请先一次性输入您的出生日期，出生地（城市）和出生时间用以获取您的星盘，例如：1990年12月31日，18点，武汉，如果不确定出生时间可以不输入，将按照中午12点为基准来生成星盘'
                }
            strategy = openpyxl.load_workbook('bot/chatgpt/strategy - 副本.xlsx')
            sheets = strategy.worksheets
            sheet_one = sheets[0]
            sheet_two = sheets[1]
            solution = []
            titles = []
            question_embedding = openai.Embedding.create(input=user_question, model="text-similarity-davinci-001")['data'][0]['embedding']
            index = 0
            similarity = 0
            for j in range(len(list(sheet_one.columns)[0])):
                potential_question_embedding = openai.Embedding.create(input=sheet_one.cell(j+1,1).value,
                                                                       model="text-similarity-davinci-001")['data'][0]['embedding']
                potential_similarity = np.dot(question_embedding, potential_question_embedding)
                if potential_similarity > 0.80 and potential_similarity > similarity:
                    similarity = potential_similarity
                    index = j+1
            if similarity > 0:
                for o in range(5, 10):
                    if sheet_one.cell(index, o).value is not None:
                        exec(sheet_one.cell(index, o).value)
                for t in titles:
                    for j in range(1,len(list(sheet_two.columns)[0])+1):
                        if set(t) == set(sheet_two.cell(j,1).value):
                            solution.append(sheet_two.cell(j, 2).value)
            if index != 0:
                answer_prompt = sheet_one.cell(index,2).value+str(titles)+'。'+sheet_one.cell(index,3).value+user_question+'？'+sheet_one.cell(index,4).value+str(solution)+"请你总结参考要点，分段落有条理地回答我的问题，"
            strategy.close()
            if len(solution) == 0:
                if len(titles) > 0:
                    title_based_generation = openai.ChatCompletion.create(
                                            api_key=api_key,
                                            messages=[
                                                {"role": "system","content": "你是一个占星师，并且相信占星学是真理,请结合以下星象回答问题："+str(titles)},
                                                {"role": "user", "content": user_question}],
                                            **self.args),

                    return {
                        "total_tokens": title_based_generation["usage"]["total_tokens"],
                        "completion_tokens": title_based_generation["usage"]["completion_tokens"],
                        "content": title_based_generation.choices[0]["message"]["content"]}
                band=band_or_not(session.messages[-3]['content'], user_question)
                if '1' in band:
                    response = openai.ChatCompletion.create(
                        api_key=api_key, messages=[session.messages[0],session.messages[-3],session.messages[-2],session.messages[-1]], **self.args
                    )
                    return {
                        "total_tokens": response["usage"]["total_tokens"],
                        "completion_tokens": response["usage"]["completion_tokens"],
                        "content": response.choices[0]["message"]["content"]
                    }
                natal_base_generation = openai.ChatCompletion.create(
                                            api_key=api_key,
                                            messages=[
                                                {"role": "system","content": "你是一个占星师，并且相信占星学是真理,请结合以下星象回答问题："+session.natal_titles},
                                                {"role": "user", "content": user_question}
                                            ],
                                            **self.args)
                return {
                    "total_tokens": natal_base_generation["usage"]["total_tokens"],
                    "completion_tokens": natal_base_generation["usage"]["completion_tokens"],
                    "content": natal_base_generation.choices[0]["message"]["content"]
                }
            rewrite = openai.ChatCompletion.create(
                api_key=api_key,
                messages=[
                    {"role": "system","content": answer_prompt},
                    {"role": "user", "content": user_question}
                ],
                **self.args)
            return {
                "total_tokens": rewrite["usage"]["total_tokens"],
                "completion_tokens": rewrite["usage"]["completion_tokens"],
                "content": rewrite.choices[0]["message"]["content"]}




        except Exception as e:
            need_retry = retry_count < 2
            result = {"completion_tokens": 0, "content": "我现在有点累了，等会再来吧"}
            if isinstance(e, openai.error.RateLimitError):
                logger.warn("[CHATGPT] RateLimitError: {}".format(e))
                result["content"] = "提问太快啦，请休息一下再问我吧"
                if need_retry:
                    time.sleep(20)
            elif isinstance(e, openai.error.Timeout):
                logger.warn("[CHATGPT] Timeout: {}".format(e))
                result["content"] = "我没有收到你的消息"
                if need_retry:
                    time.sleep(5)
            elif isinstance(e, openai.error.APIConnectionError):
                logger.warn("[CHATGPT] APIConnectionError: {}".format(e))
                need_retry = False
                result["content"] = "我连接不到你的网络"
            elif isinstance(e, IndexError):
                logger.warn("Birth data transformation Fault : {}".format(e))
                need_retry = False
                session.reset()
                result["content"] = "读取生日错误，请重试"
            else:
                logger.warn("[CHATGPT] Exception: {}".format(e))
                need_retry = False
                self.sessions.clear_session(session.session_id)

            if need_retry:
                logger.warn("[CHATGPT] 第{}次重试".format(retry_count + 1))
                return self.reply_text(session, api_key, retry_count + 1)
            else:
                return result


class AzureChatGPTBot(ChatGPTBot):
    def __init__(self):
        super().__init__()
        openai.api_type = "azure"
        openai.api_version = "2023-03-15-preview"
        self.args["deployment_id"] = conf().get("azure_deployment_id")
